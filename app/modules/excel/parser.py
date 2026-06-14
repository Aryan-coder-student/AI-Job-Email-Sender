from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from typing import Any, BinaryIO, TYPE_CHECKING

from app.core.exceptions import InvalidExcelError
from app.core.logger import get_logger
from app.modules.excel.config import ExcelParserConfig
from app.modules.excel.utils import (
    build_alias_lookup,
    build_headers,
    normalize_data,
    row_exceeds_empty_ratio,
    row_has_value,
    row_to_dict,
    select_sheets,
)
from app.modules.excel.validator import validate_excel_filename


if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

logger = get_logger(__name__)


@dataclass(frozen=True)
class ParsedExcelRow:
    sheet_name: str
    row_number: int
    data: dict[str, Any]
    normalized: dict[str, Any]

    def to_dict(self) -> dict[str, Any]:
        return {
            "sheet_name": self.sheet_name,
            "row_number": self.row_number,
            "data": self.data,
            "normalized": self.normalized,
        }


@dataclass(frozen=True)
class ParsedExcelSheet:
    name: str
    headers: list[str]
    rows: list[ParsedExcelRow]
    total_populated_rows: int

    @property
    def parsed_rows(self) -> int:
        return len(self.rows)

    def to_dict(self) -> dict[str, Any]:
        return {
            "name": self.name,
            "headers": self.headers,
            "parsed_rows": self.parsed_rows,
            "total_populated_rows": self.total_populated_rows,
            "rows": [row.to_dict() for row in self.rows],
        }


@dataclass(frozen=True)
class ParsedExcelWorkbook:
    filename: str | None
    sheets: list[ParsedExcelSheet]

    @property
    def sheet_count(self) -> int:
        return len(self.sheets)

    @property
    def total_rows(self) -> int:
        return sum(sheet.parsed_rows for sheet in self.sheets)

    @property
    def total_populated_rows(self) -> int:
        return sum(sheet.total_populated_rows for sheet in self.sheets)

    @property
    def rows(self) -> list[ParsedExcelRow]:
        return [row for sheet in self.sheets for row in sheet.rows]

    def to_dict(self) -> dict[str, Any]:
        return {
            "filename": self.filename,
            "sheet_count": self.sheet_count,
            "total_rows": self.total_rows,
            "total_populated_rows": self.total_populated_rows,
            "sheets": [sheet.to_dict() for sheet in self.sheets],
        }

    def to_company_records(self) -> list[dict[str, Any]]:
        records: list[dict[str, Any]] = []

        for row in self.rows:
            records.append(
                {
                    **row.normalized,
                    "source_sheet": row.sheet_name,
                    "source_row": row.row_number,
                    "raw_data": row.data,
                }
            )

        return records


def parse_excel_from_upload(
    file_content: bytes,
    filename: str | None = None,
    config: ExcelParserConfig | None = None,
) -> ParsedExcelWorkbook:
    """Parse Excel bytes received from an upload endpoint."""
    validate_excel_filename(filename)
    logger.info("Parsing uploaded Excel file filename=%s", filename)
    return _parse_workbook(BytesIO(file_content), filename=filename, config=config)


def _parse_workbook(
    file: BinaryIO,
    filename: str | None,
    config: ExcelParserConfig | None,
) -> ParsedExcelWorkbook:
    active_config = config or ExcelParserConfig()
    active_config.validate()

    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as error:
        raise InvalidExcelError(
            "openpyxl is required to parse Excel files. Install dependencies from requirements.txt."
        ) from error

    try:
        workbook = load_workbook(file, read_only=True, data_only=True)
    except Exception as error:
        logger.exception("Could not open Excel file filename=%s", filename)
        raise InvalidExcelError(f"Could not open Excel file: {error}") from error

    sheets = [
        _parse_sheet(sheet, active_config)
        for sheet in select_sheets(workbook.worksheets, active_config)
    ]
    workbook_result = ParsedExcelWorkbook(filename=filename, sheets=sheets)
    logger.info(
        "Parsed Excel workbook filename=%s sheets=%s rows=%s",
        filename,
        workbook_result.sheet_count,
        workbook_result.total_rows,
    )
    return workbook_result


def _parse_sheet(sheet: Worksheet, config: ExcelParserConfig) -> ParsedExcelSheet:
    header_values = next(
        sheet.iter_rows(
            min_row=config.header_row,
            max_row=config.header_row,
            values_only=True,
        ),
        None,
    )

    if not header_values or not row_has_value(header_values):
        raise InvalidExcelError(
            f"Sheet '{sheet.title}' does not contain headers on row {config.header_row}."
        )

    headers = build_headers(header_values)
    alias_lookup = build_alias_lookup(config)
    rows: list[ParsedExcelRow] = []
    total_populated_rows = 0

    for row_number, values in enumerate(
        sheet.iter_rows(min_row=config.header_row + 1, values_only=True),
        start=config.header_row + 1,
    ):
        if not row_has_value(values):
            continue

        total_populated_rows += 1

        if row_exceeds_empty_ratio(values, config.max_empty_ratio):
            continue

        if config.max_rows is not None and len(rows) >= config.max_rows:
            continue

        data = row_to_dict(headers, values)
        rows.append(
            ParsedExcelRow(
                sheet_name=sheet.title,
                row_number=row_number,
                data=data,
                normalized=normalize_data(data, alias_lookup),
            )
        )

    return ParsedExcelSheet(
        name=sheet.title,
        headers=headers,
        rows=rows,
        total_populated_rows=total_populated_rows,
    )
